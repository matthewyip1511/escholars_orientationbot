import telebot
from datetime import datetime
from telebot import types
import random
import time
import json   #to save the data from the dict
import threading
from openpyxl import load_workbook

#tele bot id
escholars = "tele_bot_id"
exco_id = "exco_id"
tracking_id = "tracking_id"

bot_id = escholars   #this id is for the testbed (now testing for ori game)
is_using_mac = False #change accordingly based on the platform used
can_change_team_name = [True]  #set to true at the start, will be locked later
is_supper_bidding = [False]    #will be True later after supper bidding starts 

#info about the game
teams = {}
stations = {}
started_users = []
registered_users = []
curr_team = []
profiles = {}   #key: telegram username item: details
admins = []   #list of users who can perform admin-related functions, more can be added later with setter functions
admins_id = []


#data for the games that are saved in the main area
latest_happenings = []  #a list to store the latest happenings
wild_cards = {}  
active_side_quests = {}
active_red_card_quests = {}
trivia_game_data = []
th_game_data = [] 
all_side_quests = {}   #key: side quest name   value: wild card that this side quest gives 


#to start the bot
bot = telebot.TeleBot(bot_id)   #initialises the bot with the telegram API
ini_time = datetime.now()      #to track when the bot started running
default_profile = {"curr": None, "team": None, "player_type" : None, "curr_profile" : None}


##########GAME DATA CLASS FOR TEAM AND STATION###########

#to contain information about each team 
class Team(object):
    def __init__(self, team_name):
        self.team_name = team_name
        self.members = []   #user id of the member
        self.points_history = []
        self.points = 0
        self.green_wildcards = {}
        self.red_wildcards = {}
        self.is_playing_t = False
        self.trivia_plays = 3   #number of times left
        self.side_quests = {}
        self.has_started_side_quest = False
        self.is_playing_sq = False
        self.green_wildcards_list = []


    def get_team_name(self):
        return self.team_name
    
    def get_team_members(self):
        return self.members

    def set_team_name(self, new_team_name):
        self.team_name = new_team_name

    def get_points(self):
        return self.points

    def add_points(self, points_value, username):
        self.points += points_value

        #log in the history for this team 
        self.points_history.append((points_value, username, form_time()))  #keeps all heppenings

        #log in the happenings!
        latest_happenings.append((points_value, username, form_time(), self.team_name))  

        #send a message to the group for points tracking! 
        msg = "(" + str(form_time()) + ") " + self.team_name + " gained " + str(points_value) + " points from " + username
        bot.send_message(tracking_id, msg)

    def add_member(self, user_id):
        self.members.append(user_id)

    def remove_member(self, user_id):
        self.members.remove(user_id)

    def is_member(self, user_id):
        return user_id in self.members   #is a member if the user id is logged as part of this team

    def get_number_members(self):
        return len(self.members)

    def get_points_history(self):
        return self.points_history
    
    def add_green_wildcard(self, name, effect):
        if name not in self.green_wildcards:
            self.green_wildcards[name] = [effect, 1]  #(effect, quantity)

        else:
            self.green_wildcards[name][1] += 1    #add one more of this wildcard

        self.green_wildcards_list.append(name)

    def get_green_wildcards(self):
        return self.green_wildcards
    
    def get_num_green_wildcards(self):
        return len(self.green_wildcards)

    def add_red_wildcard(self, name, effect):
        self.red_wildcards[name] = effect

    def get_num_red_wildcards(self):
        return len(self.red_wildcards)

    def get_red_wildcards(self):
        return self.red_wildcards

    def is_playing_trivia(self):
        return self.is_playing_t

    def set_playing_trivia(self, bool_value):
        self.is_playing_t = bool_value

    def can_play_trivia(self):
        return self.trivia_plays > 0

    def decrement_trivia(self):
        self.trivia_plays -= 1


    def is_playing_side_quest(self):
        return self.is_playing_sq
    
    def set_playing_side_quest(self, bool_value):
        self.is_playing_sq = bool_value

    def add_side_quest(self, side_quest, wild_card):
        self.side_quests[side_quest] = wild_card

    def remove_side_quest(self, side_quest):
        del self.side_quests[side_quest]

    def get_side_quests(self):
        return self.side_quests
    
    def set_side_quests(self, sq_dict):
        self.side_quests = sq_dict

    def get_remaining_side_quests(self, to_add=0):
        no_completed = 25 - len(self.side_quests) + to_add  #+1 to account for issues downstream in the adding code
        str_no_completed = str(no_completed)

        msg = str_no_completed + "/" + str(25)

        return msg

    def set_started_side_quest(self, bool_value):
        self.has_started_side_quest = bool_value

    def has_started_side_quest_check(self):
        return self.has_started_side_quest
    

    def convert_red_cards(self):
        #1. For each red card for the team, lose 20 points
        pts_per_card = -20
        num_red_wc = len(self.red_wildcards)
        pts_lose = num_red_wc * pts_per_card

        self.add_points(pts_lose, "Side Quest Admin - Red Card Penalty")

        #2. Re-set the team's red wildcards to none
        self.red_wildcards.clear()

    def remove_green_wildcard(self, wildcard):
        #1. reduce the number of this wild card by 1 first
        self.green_wildcards[wildcard][1] -= 1

        #2. if there is no more of this card then remove it entirely from the data set
        if self.green_wildcards[wildcard][1] == 0:
            del self.green_wildcards[wildcard]


    def remove_red_wildcard(self, wildcard):
        del self.red_wildcards[wildcard]


    def use_uno_reverse(self):
        if "UNO Reverse" not in self.green_wildcards:
            pass

        else:
            self.remove_green_wildcard("UNO Reverse")


    def your_loss(self):
        #1. check if the team has green wildcards
        if self.green_wildcards_list == []:
            return "NIL"
        
        else:
            #1. reduce the number of this wild card by 1 first
            latest_green_wc = self.green_wildcards_list[-1]
            self.remove_green_wildcard(latest_green_wc)
                
            return latest_green_wc

#for game station objects
class Station(object):
    def __init__(self, station_name, game_master_username, location):
        self.station_name = station_name
        self.game_master = game_master_username
        self.history = []    #to log the history of teams coming in to the station
        self.location = location

    def get_station_name(self):
        return self.station_name

    def get_game_master(self):
        return self.game_master
    
    def get_station_history(self):
        return self.history
                                   
    def add_points(self, team, pts):
        self.history.append((team, pts, form_time()))   #logs the record in a tuple of the form (team, pts, time)

    def set_location(self, location):
        self.location = location
    
    def get_location(self):
        return self.location


####################END OF DATA CLASSES#################



#READ DATA from the excel before starting 

def read_data():
    #read values from the workbook
    file_name_pc = "file_name"
    file_name_mac = "file_name"

    if is_using_mac:
        file_name = file_name_mac

    else:
        file_name = file_name_pc
        
    wb = load_workbook(filename = file_name)

    #first read for freshie
    sheet = wb["Freshie Data"]

    #get values
    r = 2  #start reading from row 2
    
    while True:
        name = sheet.cell(row=r, column=2).value
        username = sheet.cell(row=r, column=3).value
        team_name = sheet.cell(row=r, column=4).value
        player_type = sheet.cell(row=r, column=6).value
        chat_id = None

        #break conditon for eof
        if name == None:
            break

        profiles[username] = default_profile.copy()
        print(team_name)

        #depending on whether this team needs to be created or add the member to an existing team
        if team_name not in teams.keys():
            #create a new team with this name
            team_obj = Team(team_name)
            team_obj.add_member((name, username, chat_id))   #add the leader to the team as well
            teams[team_name] = team_obj
            profiles[username]["team"] = team_obj
            

        else:
            teams[team_name].add_member((name, username, chat_id))
            #print(teams[team_name].members)
            profiles[username]["team"] = teams[team_name]

        #rest of the set up for profile    
        profiles[username]["player_type"] = player_type
        profiles[username]["name"] = name

        #if the player is a Exco or Game Master give him admin rights
        if player_type == "Exco" or player_type == "Game Master":
            admins.append(username)

        #once done reading this guy move to the next row
        r += 1
        #print(profiles)


    ##now read for station masters
    r = 2  #start reading from row 2
    sheet = wb["Station Data"]
    
    while True:
        station_name = sheet.cell(row=r, column=2).value
        station_master = sheet.cell(row=r, column=3).value
        username = sheet.cell(row=r, column=4).value
        print(username)
        location = sheet.cell(row=r, column=5).value

        #break conditon for eof
        if station_name == None:
            break

        #create a new station 
        stn = Station(station_name, username, location)
        stations[station_name] = stn  #set this station and its name in the stations dict

        #add the station to the person's profile
        profiles[username]["station"] = stn

        #once done reading this guy move to the next row
        r += 1
        #print(profiles)


    ##last step, read for exco
    r = 2
    sheet = wb["Exco Data"]

    while True:
        name = sheet.cell(row=r, column=2).value
        username = sheet.cell(row=r, column=3).value
        chat_id = None

        #break conditon for eof
        if name == None:
            break

        profiles[username] = default_profile.copy()
        profiles[username]["player_type"] = "Exco"
        admins.append(username)

        r += 1

    #READ DATA FOR THE GAMES
        
    #1. read data for the trivia game
    r = 3
    sheet = wb["Trivia Game Data (Raw)"]

    #get the data for the trivia game and add to the list for the trivia games data
    while True:
        qn = sheet.cell(row=r, column=2).value

        #check eof
        if qn == None:
            break
        
        correct_ans = sheet.cell(row=r, column=3).value
        op2 = sheet.cell(row=r, column=4).value
        op3 = sheet.cell(row=r, column=5).value
        op4 = sheet.cell(row=r, column=6).value

        trivia_game_data.append((qn, correct_ans, op2, op3, op4))

        r += 1

    #2. Get data for the treasure hunt
    r = 2   #start reading from second row
    sheet = wb["Treasure Hunt Data"]

    while True:
        clue = sheet.cell(row=r, column=2).value
        
        #check eof
        if clue == None:
            break
        
        ans = sheet.cell(row=r, column=3).value

        th_game_data.append((r-1, clue, ans))   #r-1 to indicate the station number (one less than the row) 

        r += 1
        

    #read data for the wild cards for the side quesst
    r = 2   #start reading from second row
    sheet = wb["Wild Cards"]

    while True:
        card_name = sheet.cell(row=r, column=3).value
        card_effect = sheet.cell(row=r, column=4).value
        
        #check eof
        if card_name == None:
            break

        #add to the data
        wild_cards[card_name] = card_effect

        r += 1

    print(wild_cards)


    #then read for the side quest and its corresponding wild card

    r = 2   #start reading from second row
    sheet = wb["Side Quests"]

    while True:
        side_quest_name = sheet.cell(row=r, column=1).value
        wild_card = sheet.cell(row=r, column=3).value
        colour = sheet.cell(row=r, column=4).value
        description = sheet.cell(row=r, column=5).value
        
        #check eof
        if side_quest_name == None:
            break

        #add to the data
        all_side_quests[side_quest_name] = (wild_card, colour, description)

        r += 1

    #load side quest data to all teams
    for t in teams.values():
        t.set_side_quests(all_side_quests.copy())



    print("finished reading")

#run this at the start to read data off the excel sheet
read_data()


##impt to save data before refreshing the files to check for bugs

@bot.message_handler(commands=['admin_save'])

def admin_save(message):

    with open("C:/users/matth/ori_data.json", "w") as outfile:
        json.dump(profiles, outfile)

    bot.send_message(message.from_user.id, "Information successfully saved!")

    

########START OF THE MAIN CODE WITH THE START FUNCTION#####################
@bot.message_handler(commands=['start'])

def send_welcome(message):
    username = message.from_user.username
    user_id = message.from_user.id
    print(user_id)

    #add the user to a list of started users, that would be able to receive broadcasts later
    if user_id not in started_users:
        started_users.append(user_id)

    #check if there is such a profile for the user.
    #if the user already has a profile, let them know what they are and what team they are in (if applicable)
    if username in profiles.keys():     #user already has an account
        player_type = profiles[username]["player_type"]   #check the player type from profile first
        msg = ""   #msg is empty first
        print(player_type)

        team_name = profiles[username]["team"].get_team_name()

        

        if player_type == "Freshie":
            msg += "Welcome! You are joined as a <b>Freshie</b> in team <u>" + team_name + "</u>!"

        elif player_type == "Game Master":
            station_name = profiles[username]["station"].get_station_name()
            msg += "Welcome! You are joined as a <b>Game Master</b> in <u>" + station_name + "</u> and tagged to team <u>" + team_name + "</u>!"

        elif player_type == "OGL":
            msg += "Welcome! You are joined as a <b>OGL</b> in <u>" + team_name + "</u>!"

        elif player_type == "Exco":
            msg += "Welcome! You are joined as an <b>Exco</b>, tagged to team <u>" + team_name  + "</u>!"


        #suffix for all users to be added to the msg after distinguishing
        msg += "\n\nFor more information about your OG, press /myteam.\nFor other information about the orientation, press /information\n\nIf you think that this may be a mistake, contact Matthew Yip @matthewyip"

        bot.send_message(message.from_user.id, msg, reply_markup=remove_keyboard(), parse_mode="HTML")   #send that message to the user and then end the function there
        return

        
 
    #for non-signed up users - move to this page
    welcome_message = "Welcome to the 2023 E-scholars Orientation Bot! Your account is not currently registered in the bot, please contact Matthew @matthewyip for assistance"

    profiles[username] = default_profile.copy()  #create a profile for this new user first 

    bot.send_message(message.from_user.id, welcome_message)   #message sent to the user to inform them that the system did not capture them 


#for those who are not registered in the bot
@bot.message_handler(commands=["jointeam"])

def jointeam(message):
    #check status first. if the guy has a team then do not allow
    username = message.from_user.username
    if profiles[username]["team"] is not None:
        bot.send_message(message.from_user.id, "Sorry! This function is only for those who are not assigned a team")
        return
    

    keyboard = generate_keyboard(*teams.keys())

    bot.send_message(message.from_user.id, "Which team are you joining?", reply_markup=keyboard)
    bot.register_next_step_handler(message, actl_jointeam)

def actl_jointeam(message):
    #data
    username = message.from_user.username
    fullname = message.from_user.full_name 
    team_name = message.text
    team = teams[team_name]

    team.add_member((fullname, username, None))
    profiles[username]["team"] = team

    bot.send_message(message.from_user.id, "Done! You are now a member of team " + team_name, reply_markup=remove_keyboard())

@bot.message_handler(commands=["changeteam"])

def call_changeteam(message):
    username = message.from_user.username
    user_team_name = profiles[username]["team"].get_team_name()

    keyboard = generate_keyboard(list(filter(lambda team: team != user_team_name, teams.keys())))

    bot.send_message(message.from_user.id, "Ok, which team would you like to change to?", reply_markup=keyboard)



@bot.message_handler(commands=["help"])

def send_help(message):
    help_msg = "This bot will be used for tracking points and team scores throughout the orientaion. If you face any technical difficulties,\
                do not hesitate to contact @matthewyip on telegram!"

    bot.send_message(message.from_user.id, help_msg)
    print("help answered")

#check how many teams have registered 

@bot.message_handler(commands=['checkteams'])

def check_teams(message):
    user_username = message.from_user.username

    no = len(teams)

    msg = "Hello! There are currently " + str(no) + " teams in the game. They are:\n"

    msg += form_teams(teams)   #convert the list of teams to word outputable format

    msg += "\n\n"
    msg += "Here's the latest updates in the orientation:\n\n"

    n = min(5, len(latest_happenings))
    print(n)
    for pts, username, time, team_name in latest_happenings[:-n-1:-1]:
        line = "(" + time + ") " + team_name + " gained " + str(pts) + " points from " + username + "!\n"
        msg += line

    
    #FOR ADMINS add the function to see data for any team
    if is_admin(user_username):
        msg += "\n\n(For Admins) To know more about a specific team, press /checkteams_details"


    bot.send_message(message.from_user.id, msg)   #print out the number of teams, and the team names


@bot.message_handler(commands=["checkteams_details"]) 

def checkteams_details(message):
    #check admin status
    username = message.from_user.username
    if not is_admin(username):
        say_no()
        return

    #otherwise ask for team
    keyboard = generate_keyboard(*list(teams.keys())) #keyboard with names of teams

    bot.send_message(message.from_user.id, "Hello! Which team would you like to know more about?", reply_markup=keyboard)
    bot.register_next_step_handler(message, show_team_details)


def show_team_details(message):
    #data
    username = message.from_user.username
    team_name = message.text
    team = teams[team_name]

    #get info for this team 
    give_team_information(username, team, message)

@bot.message_handler(commands=["full_points_history"])

def full_points_history(message):
    msg = "Ok, here is the full list of points information:\n\n"
    
    for pts, username, time, team_name in latest_happenings[::-1]:
        line = "(" + time + ") " + team_name + " gained " + str(pts) + " points from " + username + "!\n"
        msg += line

    bot.send_message(message.from_user.id, msg)


@bot.message_handler(commands=["checkuser"])

def call_check_user(message):
    #check if the user is an admin first and allowed to use function
    if not is_admin(message.from_user.username):
        say_no()
        return
    
    #otherwise allow this
    bot.send_message(message.from_user.id, "Ok, please send the telegram handle of the user you are checking")
    bot.register_next_step_handler(message, check_user)

def check_user(message):

    user_username = message.text
    user_name = profiles[user_username]["name"]

    #check first if there is such a user
    if not user_username in profiles:
        bot.send_message(message.from_user.id, "Sorry! There is no such user registered in the orientation")
        return
    
    else:
        print(profiles[user_username])
        team_name = profiles[user_username]["team"].get_team_name()
        
        msg = user_name + " @" + user_username + " is a member of team " + team_name
        bot.send_message(message.from_user.id, msg)
    

@bot.message_handler(commands=["checkstations"])

def call_check_stations(message):
    #data
    username = message.from_user.username

    #get the list of stations
    stns_list = list(stations.keys())

    #1. message header
    msg = "Hello, here are the Station Games in this orientation:\n\n"

    #2. add the data
    for stn in stns_list:
        msg += stn + "\n"

    #3. ask them if they want to know more
    msg += "\nTo know more about a particular station, press /checkstations_details"

    #send it out 
    bot.send_message(message.from_user.id, msg)


@bot.message_handler(commands=["checkstations_details"])

def check_stations_details(message):
    stns_list = list(stations.keys())
    keyboard = generate_keyboard(*stns_list)

    #1. form header
    msg = "Which Station would you like to find out more about?"

    bot.send_message(message.from_user.id, msg, reply_markup = keyboard)
    bot.register_next_step_handler(message, give_stations_details)


def give_stations_details(message):
    #data
    stn_name = message.text
    stn = stations[stn_name]
    game_master_username = stn.get_game_master()
    location = stn.get_location()
    game_master_name = profiles[game_master_username]["name"]

    #msg header
    msg = "Ok, here is the information for this Station Game:\n\n"

    msg += "Station Name: " + stn_name + "\n"
    msg += "Game Master Name: " + game_master_name + " @" + game_master_username + "\n"
    msg += "Location: " + location

    bot.send_message(message.from_user.id, msg)

#function to add points to the team (for station admin only)
@bot.message_handler(commands=["addpoints"])

def add_points(message):
    #admin check
    if not is_admin(message.from_user.username):
        say_no()
        return
    
    #if there is a profile
    username = message.from_user.username
    if message.text != "/addpoints":
        profiles[username]["curr_profile"] = message.text

    msg = "There are the following teams in the game:\n"

    msg += form_teams(teams)

    msg += "\n\nWhich team do you want to add to?\n(To cancel this operation, press Cancel)"

    # Create a custom keyboard
    keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    
    # Add buttons to the keyboard
    teams_list = list(teams.items())
    teams_list.sort(key=lambda team: team[1].get_points(), reverse=True)
    
    for team_name, team in teams_list:
        button = types.KeyboardButton(team_name)
        keyboard.add(button)

    keyboard.add("Cancel")


    bot.send_message(message.from_user.id, msg, reply_markup=keyboard)

    #then wait for the reply for which team they want to add to
    bot.register_next_step_handler(message, ask_points)

def ask_points(message):
    #check for cancel
    if message.text == "Cancel":
        bot.send_message(message.from_user.id, "Ok, this process has been aborted", reply_markup=remove_keyboard())
        return
    
    #if not process the rest of the code
    #data 
    team_name = message.text
    username = message.from_user.username
    
    profiles[username]["curr"] = team_name
    bot.send_message(message.from_user.id, "OK, how many points do you want to add?", reply_markup=remove_keyboard())

    #then wait for the input on how many points they want added to the team
    bot.register_next_step_handler(message, confirm_entry)

def confirm_entry(message):
    #get data 
    pts = int(message.text)
    username = message.from_user.username
    team_name = profiles[username]["curr"]
    profiles[username]["curr_pts"] = pts

    #confirm with them 
    msg = "You are adding " + str(pts) + " points to team " + team_name + "."
    msg += "\nTo confirm, press Yes. To start over, press Cancel"
    keyboard = generate_keyboard("Yes", "Cancel")

    bot.send_message(message.from_user.id, msg, reply_markup=keyboard)
    bot.register_next_step_handler(message, process_choice)

def process_choice(message):
    #see what the choice is
    choice = message.text

    if choice == "Yes":
        actl_add_to_team(message)

    elif choice == "Cancel":
        add_points(message)


def actl_add_to_team(message):
    #get the information
    username = message.from_user.username
    pts = profiles[username]["curr_pts"]
    print(username)
    team_name = profiles[username]["curr"]
    team = teams[team_name]

    #perform the addition
    prev_points = team.get_points()

    #check who is adding the points
    player_type = profiles[username]["player_type"]

    if profiles[username]["curr_profile"] != None:
        name = profiles[username]["curr_profile"]

    elif player_type == "Game Master":
        name = profiles[username]["station"].get_station_name() 

    elif player_type == "Exco":
        name = username + " - Exco"

    #add points to this team and check how many points they now have 
    team.add_points(pts, name)
    new_points = team.get_points()

    if new_points >= prev_points:
        msg = "Done! Team " + team_name + " now has " + str(new_points) + " points, up from " + str(prev_points)

    else:
        msg = "Done! Team " + team_name + " now has " + str(new_points) + " points, down from " + str(prev_points)

    bot.send_message(message.from_user.id, msg, reply_markup=remove_keyboard())


##alternate add points for when you need to use a diff profile###

@bot.message_handler(commands=["addpoints_as_profile"])

def add_points_as_profile(message):
    if not is_admin(message.from_user.username):
        say_no()
        return
    
    keyboard = generate_keyboard("Water Games", "Supper Bidding")    #to give two possible profiles for them to use as

    bot.send_message(message.from_user.id, "Ok, please indicate the profile which you would like to add points as (You can also type a custom profile other than the custom buttons given)", reply_markup=keyboard)
    bot.register_next_step_handler(message, add_points)
    
#auxillary functions for later 
@bot.message_handler(commands=['timenow'])

def send_curr_time(message):
    curr_time = datetime.now()
    bot.send_message(message.from_user.id, curr_time)

@bot.message_handler(commands=['startime'])

def send_start_time(message):
    bot.send_message(message.from_user.id, ini_time)

@bot.message_handler(commands=['checkadmin'])

def check_admin(message):

    if (message.from_user.username in admins):
        msg = "Yep, you are an admin with admin rights!"

    else:
        msg = "Nope, you are not an admin :("

    bot.send_message(message.from_user.id, msg)


#########Enquiry############################

@bot.message_handler(commands=['myteam'])

def myteam(message):
    #find out the team information first
    username = message.from_user.username
    team = profiles[username]["team"]

    #execute the code
    give_team_information(username, team, message)


def give_team_information(username, team, message):
    #find out the team information first
    team_name = team.get_team_name()

    #build the message with the team information 
    msg = "Welcome! Your team information is as follows:\n\n"
    msg += "Team Name: " + team_name + "\n"
    msg += "Points = " + str(team.get_points()) +"\n\n"

    msg += "Number of members = " + str(team.get_number_members()) + "\n\n"
    msg += "<u>Members:</u>\n\n"

    for tup in team.get_team_members():
        name, name_username = tup[0], tup[1]
        msg += name + " @" + name_username + "\n"


    msg += "\n<u>Wild Cards:</u>\n\n"

    msg += "Green Wild Cards = " + str(team.get_num_green_wildcards()) + "\n\n"

    #display green wild cards
    for wc_name, tup in list(team.get_green_wildcards().items()):
        num = tup[1]
        msg += str(num) + "x "   #eg. 1x Item Revealer 
        msg += wc_name
        msg += "\n"

    msg += "\nRed Wild Cards = " + str(team.get_num_red_wildcards()) + "\n\n"

    #display red wild cards
    for key in list(team.get_red_wildcards().keys()):
        msg += "1x "
        msg += key
        msg += "\n"

    msg += "\nTo use up any of the red wild cards at any time, press /use_red_wildcard"
    msg += "\n\nTo know more about the wildcards you are holding on to, press /wildcards_information"
    
    #only add renaming team if its allowed
    if can_change_team_name[0]:
        msg += "\n\nIf you would like to re-name the team, press /rename_team"
        
    msg += "\n\n\n<b>Points History (Last 5 entries)</b>:\n\n"


    #add in information for the points history of the team
    pts_history = team.get_points_history()[::-1]

    no_entries = min(5, len(pts_history))  #whichever is lower

    for i in range(no_entries):
        points, username, time = pts_history[i]
        
        msg += "(" + time + ") " + str(points) + " points earned from " + username + "!\n"   #to state where points earned / how much 

    #end by telling the user that there is an option to check the full list should they wish to
    msg += "\nTo see the full list of points history, press /myteam_points"
    bot.send_message(message.from_user.id, msg, parse_mode='HTML', reply_markup=remove_keyboard())


@bot.message_handler(commands=['use_red_wildcard'])  

def call_use_red_wildcard(message):
    #data
    username = message.from_user.username
    team = profiles[username]["team"]
    team_name = team.get_team_name()
    
    red_wildcards = list(team.get_red_wildcards().keys())
    print(red_wildcards)

    keyboard = generate_keyboard(*red_wildcards)

    bot.send_message(message.from_user.id, "Ok, which red wildcard would you like to use?", reply_markup=keyboard)
    bot.register_next_step_handler(message, use_red_wildcard)

def use_red_wildcard(message):
    #data
    username = message.from_user.username
    team = profiles[username]["team"]
    red_wildcard = message.text
    profiles[username]["red_wildcard"] = red_wildcard
    effect = wild_cards[red_wildcard]

    #1. remind them what this wildcard is asking for
    msg = red_wildcard + ": " + effect
    bot.send_message(message.from_user.id, msg)

    #2. let them know of the process  
    bot.send_message(message.from_user.id, "Please go ahead and send the photo / video as listed by the instructions above", reply_markup=remove_keyboard())
    bot.register_next_step_handler(message, log_red_wildcard)


def log_red_wildcard(message):
    #get data
    username = message.from_user.username
    red_wc = profiles[username]["red_wildcard"]
    team = profiles[username]["team"]
    team_name = team.get_team_name()
    sender_id = message.from_user.id   #get the chat id of the sender to get back to them later

    #1.to the exco group, send the opener first (tell them which team sent in and which sq they are attempting)
    first_msg = "Submission for red card from " + team_name + ":"
    first_msg += "\nRed Card: " + red_wc

    obj = wild_cards[red_wc]
    first_msg += "\n\nObjective: " + obj

    bot.send_message(exco_id, first_msg)

        
    #2. process what the user sent to the bot

    if message.content_type == 'text':
        #if just a text, simple just send it over to the admin chat
        to_send = message.text

        #broadcast the message to all started users 
        bot.send_message(exco_id, to_send)

    elif message.content_type == 'photo':
        #define the photo
        photo_id = message.photo[-1].file_id
        
        # Send the photo to the exco chat
        bot.send_photo(exco_id, photo_id)

    elif message.content_type == "video":
        #define the video
        video_id = message.video.file_id

        #send video
        bot.send_video(exco_id, video_id)
        
        #to include support for videos to be sent to the bot later


    #3. close with the last msg 
    last_msg = "To approve / reject red cards answers at any time, press /approve_red_wildcard"
    msg = bot.send_message(exco_id, last_msg)
    msg_id = msg.message_id
    
    #4. save this info to the log first to be retrieved later 
    active_red_card_quests[team_name + ", " + red_wc] = (red_wc, team_name, msg_id, sender_id)   #key is team_name, sq, and the value is the id of the msg and sender

    #5. then at the end tell the user that the photo / text was sent out successfully
    bot.send_message(message.from_user.id, "Your photo/video has been successfully sent! You will be notifed again after an \
admin reviews your submission")

    
#approve side quest and give out wildcards function, only to be used by admins
@bot.message_handler(commands=["approve_red_wildcard"])

def call_approve_red_wildcard(message):
    username = message.from_user.username
    profiles[username]["to_delete_rc"] = []
    profiles[username]["to_delete_rc"].append(message.message_id)

    #get the data of the submitted side quests first
    active_rc_items = active_red_card_quests.keys()  #convert the list of tuples to team, sq
    print(active_rc_items)
    keyboard = generate_keyboard(*active_rc_items)

    #then sent this to the user and ask which one they want to approve

    msg = bot.reply_to(message, "Which red card would you like to judge?", reply_markup=keyboard)

    #make a list of the messages to be deleted later 
    msg_id = msg.message_id
    profiles[username]["to_delete_rc"].append(msg_id)

    #next step
    bot.register_next_step_handler(message, red_card_yes_or_no)

def red_card_yes_or_no(message):
    #register which team / side quest you are looking at 
    username = message.from_user.username
    curr_sq = message.text   #the text of the side quest
    profiles[username]["curr"] = curr_sq   

    #log data of the messages to delete
    profiles[username]["to_delete_rc"].append(message.message_id)

    #give them the value as required

    #ask if it is a yes or no
    keyboard = generate_keyboard("Yes", "No")  #keyboard has two options
    msg = bot.reply_to(message, "Is this correct?", reply_markup=keyboard)
    msg_id = msg.message_id
    profiles[username]["to_delete_rc"].append(msg_id)

    bot.register_next_step_handler(message, red_card_check_answer)

def red_card_check_answer(message):
    #log data
    username = message.from_user.username
    profiles[username]["to_delete_rc"].append(message.message_id)

    #get the ans 
    ans = message.text

    #process
    if ans == "Yes":
        #data
        full_name = message.from_user.full_name
        rc_text = profiles[username]["curr"]
        rc_name, team_name, msg_id, sender_id = active_red_card_quests[rc_text]
        team = teams[team_name]

        #1. tell the team they got it
        msg = "Congrats! Your submission for " + rc_name + " is correct and the red card has been removed from your team"
        bot.send_message(sender_id, msg)
  
        #2. add the wildcard to the team and remove from team list (and effects if any)
        team.remove_red_wildcard(rc_name)
        
        #3. edit the original message in the group to say that this has been ammended 
        bot.edit_message_text(chat_id=exco_id, message_id=msg_id, text="Edit: Approved by " + full_name + " @" + username + " at " + form_time() + "H")

        #eg. the message is now "Edit: Approved by Matthew Yip @matthewyip at 1000H"


        #4. tell the group it has been awarded
        bot.reply_to(message, "Great! Red Card " + rc_name + " has been removed from team " + team_name, reply_markup=remove_keyboard())


        #5. Update the data to remove this entry
        active_red_card_quests.pop(rc_text)


        #6. Delete the back and forth messages for it to be neater
        for msg_id in profiles[username]["to_delete_rc"]:
            bot.delete_message(chat_id=exco_id, message_id=msg_id)


    elif ans == "No":
        #reply msg and log the msg sent
        msg = bot.reply_to(message, "Ok, plese include a message to be sent to the OG to inform them why the submission is not correct", reply_markup=remove_keyboard())
        msg_id = msg.message_id
        profiles[username]["to_delete_rc"].append(msg_id)

        bot.register_next_step_handler(message, red_cards_tell_them_no)

    else:
        bot.reply_to(message, "Sorry, that is not a valid option! Please press the buttons for Yes or No")
        bot.register_next_step_handler(message, side_quest_check_answer)

def red_cards_tell_them_no(message):
    #log data
    username = message.from_user.username
    profiles[username]["to_delete_rc"].append(message.message_id)

    #get the text to send them 
    rejection_reason = message.text

    #get data
    username = message.from_user.username
    full_name = message.from_user.full_name
    rc_text = profiles[username]["curr"]
    rc_name, team_name, msg_id, sender_id = active_red_card_quests[rc_text] 
    team = teams[team_name]
    
    #1. Tell the team that the submission is not correct
    bot.send_message(sender_id, "Hey, that was not correct :( Here's a message from the admins:\n\n" + rejection_reason +"\n\nTo re-submit at any time, press /use_red_wildcard")


    #2. Edit the message to show that it is not correct
    new_text = "Edit: Rejected by " + full_name + " @" + username + " at " + form_time() + "H\n\n"
    new_text += "Reason: " + rejection_reason
    bot.edit_message_text(chat_id=exco_id, message_id=msg_id, text=new_text)


    #3. Send in the exco chat that it is rejected
    bot.send_message(exco_id, "Got it, informed the team that it is not correct")

    #4. Remove information from active, and set the team to not playing 
    active_red_card_quests.pop(rc_text)

    #5. Delete all texts for neatness
    for msg_id in profiles[username]["to_delete_rc"]:
        bot.delete_message(chat_id=exco_id, message_id=msg_id)




@bot.message_handler(commands=['myteam_points'])

def give_points_history(message):
    #to show the points for ENTIRE history
    username = message.from_user.username
    team = profiles[username]["team"]
    pts_history = team.get_points_history()

    msg = "This is the full list of points history for your OG:\n\n"
    
    for i in range(len(pts_history)):
        points, username, time = pts_history[i]
        
        msg += "(" + time + ") " + str(points) + " points earned from " + username + "!\n"   #to state where points earned / how much

    bot.send_message(message.from_user.id, msg)

@bot.message_handler(commands=['rename_team'])

def call_rename_team(message):
    #check if you can still change name
    if not can_change_team_name[0]:
        bot.send_message(message.from_user.id, "Sorry! You can no longer change your team name after the start of the orientation :(")

    else:
        bot.send_message(message.from_user.id, "Ok, what name do you want to change your team name to?")

        bot.register_next_step_handler(message, rename_team)
    

def rename_team(message):

    #find out the team information first
    username = message.from_user.username
    old_team_name = profiles[username]["team"].get_team_name()
    team = teams[old_team_name]
    og_number = old_team_name[:4]   #first 4 characters is the og number 


    new_team_name = og_number + " - " + message.text

    team.set_team_name(new_team_name)  #change the team name first
    teams[new_team_name] = team   #set the team with the new key in the teams dict

    del teams[old_team_name]   #remove the old entry

    #let the user know that the thing is done
    bot.send_message(message.from_user.id, "Done! Your team name has been successfully changed to " + new_team_name)

@bot.message_handler(commands=["wildcards_information"])

def green_wildcards_information(message):
    #data
    username = message.from_user.username
    team = profiles[username]["team"]
    team_name = team.get_team_name()

    #form the message to send to the user
    msg = "Ok, here is the information for the Wild Cards you are holding on to:\n\n"

    green_wc = team.get_green_wildcards().keys() 
    red_wc = team.get_red_wildcards().keys()

    msg += "<u>Green Wild Cards:</u>\n\n"
    for key in green_wc:
        effect = wild_cards[key]
        msg += key + ": " + effect + "\n\n"

    msg += "\n<u>Red Wild Cards</u>:\n\n"
    for key in red_wc:
        effect = wild_cards[key]
        msg += key + ": " + effect + "\n"

    bot.send_message(message.from_user.id, msg, parse_mode="HTML")

@bot.message_handler(commands=["lock_rename_team"])

def lock_change_team_name(message):
    #convert to the other boolean
    can_change_team_name[0] = not can_change_team_name[0]

    can_onot = "can" if can_change_team_name[0] else "cannot"
    
    #send new status
    bot.send_message(message.from_user.id, "Ok, teams now " + can_onot + " change their team names")
    

@bot.message_handler(commands=["mystation"])

def call_my_station(message):
    username = message.from_user.username
    if profiles[username]["player_type"] != "Game Master":
        bot.send_message(message.from_user.id, "Sorry, this function is for game masters only :(")
        return

    #if can show them the data
    stn = profiles[username]["station"]
    stn_name = stn.get_station_name() 
    location = stn.get_location()


    #1. Message header
    msg = "Hello! Here's the information for your station game:\n\n"


    #2. add the details
    msg += "Station Game Name: " + stn_name + "\n"
    msg += "Location: " + location + "\n\n"


    #3. show points history for this station 
    msg += "Here is the points history for this station:"

    
    #now send it out
    bot.send_message(message.from_user.id, msg)
    

##################GAMES SECTION################################

##GAME SETUP: FUNCTION TO SUPPORT GAME STARTING

@bot.message_handler(commands=["start_games"])

def show_games_commands(message):
    msg = "Welcome! Here are the list of games we are using in this orientation:\n\n"

    msg += "1. Treasure Hunt /start_game_th"
    msg += "2. Bingo /start_game_bingo"

    bot.send_message(message.from_user.id, msg)

    

#GAME1: this is the trivia game 
@bot.message_handler(commands=["start_game_trivia"])

def start_game_trivia(message):
    #check if anyone else in the team is currently playing (team should only play at one phone at one time)(only for freshies)
    username = message.from_user.username
    can_play = True

    if profiles[username]["player_type"] == "Freshie":
        if profiles[username]["team"].is_playing_trivia():
            bot.send_message(message.from_user.id, "Sorry! Your teammate is currently playing the game, only one OGL from each OG can play the game at one time!")
            can_play = False

        elif not profiles[username]["team"].can_play_trivia():
            bot.send_message(message.from_user.id, "Your OG has reached the limit of 3 points-earning trivia games to be played during the orientation! You may still play the game but will not be able to earn any more points.")
            

    else:
        profiles[username]["team"].set_playing_trivia(True)
        
    #code proper 
    if can_play:   #if cannot play these lines of code would not run 
    
        #shuffle the questions, and make a copy for this player
        profiles[username]["trivia_game_data"] = trivia_game_data.copy()
        random.shuffle(profiles[username]["trivia_game_data"])
    
        #keyboard for custom
        keyboard = generate_keyboard("Start", "Cancel")

        #set up data for user
        profiles[username]["rounds"] = [None, None, None, None]  #[total, current, remaining, points]


        msg = "Welcome to the trivia game! Your OG can attempt this game at any time (up to 3 rounds with points!) as long as you are free :)\n\n"

        msg += "The rules of the game are as follows:\n\nThere are <b>10</b> rounds in each game. In each round, there will be a question and 4 options for the answer. Pick the answer that you think is correct. 1 point \
will be awarded for the correct answer while no points will be awarded / deducted for the wrong answer. But be fast! You will only have <b>10s</b> to answer each question."

        msg += "\n\nTo begin, press Start. To end this operation and play at a later time, press Cancel"

    
        bot.send_message(message.from_user.id, msg, reply_markup=keyboard, parse_mode='HTML')

        bot.register_next_step_handler(message, game_trivia_rounds_check)


def game_trivia_rounds_check(message):
    #check if to cancel or proceed. if cancel then end the function here
    if message.text == "Cancel":
        username = message.from_user.username
        profiles[username]["team"].set_playing_trivia(False)
        bot.send_message(message.from_user.id, "Ok, this process has been aborted. You may start the game at any time by pressing /start_game_trivia", reply_markup=remove_keyboard())
        return
    
    #set the rounds to play and send out a message
    r = 10
    username = message.from_user.username
    profiles[username]["rounds"] = [r, 1, r, 0]
    profiles[username]["timed_out"] = True   #set as True first as the default

    keyboard = generate_keyboard("Start")

    bot.send_message(message.from_user.id, "Great! Press start to start the game of 10 rounds", reply_markup=keyboard)
    bot.register_next_step_handler(message, play_game_trivia)

    #print(trivia_game_data)
    

def play_game_trivia(message):
    username = message.from_user.username
    total, current, remaining, pts = profiles[username]["rounds"]

    if False:
        pass

    else:
        msg = "Question for this round " + str(current) + "/" + str(total) + ":\n\n"

        question = profiles[username]["trivia_game_data"][current - 1]
        qn, correct_ans, op2, op3, op4 = question  #get out the individual values
        
        msg += qn
        msg += "\n\n"

        options = [correct_ans, op2, op3, op4]
        random.shuffle(options)

        for op in options:
            msg += str(op)
            msg += "\n"

        keyboard = generate_keyboard(*options)

        bot.send_message(message.from_user.id, msg, reply_markup=keyboard)
        profiles[username]["timed_out"] = True  #set back to True each round


        bot.register_next_step_handler(message, game_trivia_round)

        #send a reminder after 5s
        time.sleep(5)

        if profiles[username]["timed_out"] and profiles[username]["rounds"][1] == current:   #to see if its still the same question
            bot.send_message(message.from_user.id, "5 Seconds Remaining!", reply_markup=keyboard)

        else:
            return
        
        #time out the question and decrement by 1
        time.sleep(5)

        if profiles[username]["timed_out"] and profiles[username]["rounds"][1] == current:   #to see if its still the same question
            msg = "Sorry! The round has timed out, press to continue to the next round."
        
            profiles[username]["rounds"][1] += 1
            profiles[username]["rounds"][2] -= 1

            bot.send_message(message.from_user.id, msg , reply_markup=generate_keyboard("Moving on!"))
            bot.register_next_step_handler(message, play_game_trivia)
            


def game_trivia_round(message):
    username = message.from_user.username
    total, current, remaining, pts = profiles[username]["rounds"]
    profiles[username]["timed_out"] = False     #user did not time out if he made it to this part
    
    question = profiles[username]["trivia_game_data"][current - 1]
    qn, correct_ans, op2, op3, op4 = question[0], str(question[1]), str(question[2]), str(question[3]), str(question[4])  #get out the individual values
    keyboard = generate_keyboard("Moving on!")

    if message.text == "Moving on!":
        return

    elif message.text == correct_ans:
        bot.send_message(message.from_user.id, "Congratulations! That is the correct answer, your team just earned 1 point! Now on to the next question!", reply_markup=keyboard)
        profiles[username]["rounds"][3] += 1

    else:
        bot.send_message(message.from_user.id, "Aww, that was not the right answer :( But it's ok, on to the next question!", reply_markup=keyboard)
        
    profiles[username]["rounds"][1] += 1
    profiles[username]["rounds"][2] -= 1

    #check if we reached end of game (less than 0 if the game is bugged)
    total, current, remaining, pts = profiles[username]["rounds"]  #update the data
    if remaining <= 0:
        bot.send_message(message.from_user.id, "That's the end of the game, your team earned " + str(pts) + " points! You may start a new game anytime by pressing /start_game_trivia", reply_markup=remove_keyboard())

        #add points if the person is a Freshie
        if True: #profiles[username]["player_type"] == "Freshie":  for now change later
            profiles[username]["team"].set_playing_trivia(False)
            profiles[username]["team"].add_points(pts, "Trivia Game")
            profiles[username]["team"].decrement_trivia()
        
    #otherwise continue
    else:
        bot.register_next_step_handler(message, play_game_trivia)


    
def generate_keyboard(*values):
    keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)

    for value in values:
        button = types.KeyboardButton(value)
        keyboard.add(button)

    return keyboard

def generate_inline_keyboard(*values):
    #for inline keyboard
    keyboard = types.InlineKeyboardMarkup()

    for value in values:
        button = types.InlineKeyboardButton(value, callback_data=value)
        keyboard.add(button)


    return keyboard


#GAME2: game for the treasure hunt

@bot.message_handler(commands=["start_game_treasure_hunt"])


def start_game_treasure_hunt(message):
    username = message.from_user.username
    
    #check if the user has started game
    if profiles[username]["team"].is_playing_treasure_hunt():
        bot.send_message(message.from_user.id, "Welcome back! Your team has already started the game, you may send a photo for the station at any time by pressing /th_send_answer")

    else:
        msg = "Welcome to the treasure hunt! You have now started the game and can send in answers anytime. This is how the bot will work:\n\n"

        msg += "There are a total of 10 stations in the treasure hunt which you can complete in any order. To send an answer, press /th_send_answer"
        
        bot.send_message(message.from_user.id, msg)
        profiles[username]["team"].set_playing_treasure_hunt(True)  #set status for the whole team

    


@bot.message_handler(commands=["th_send_answer"])

def th_send_photo(message):
    username = message.from_user.username
    team = profiles[username]["team"]
    th_remaining = team.get_th_remaining()

    #if the team has not started ask them to start first

    if not profiles[username]["team"].is_playing_treasure_hunt():
        bot.send_message(message.from_user.id, "Your team has not yet started the treasure hunt game! Press /start_game_treasure_hunt to start the game first")

    
    else:
        keyboard = generate_keyboard(*th_remaining)
        #need to add a way to show which stations are remaining, and the completion status
        bot.send_message(message.from_user.id, "Which station would you like to send for?", reply_markup=keyboard)
        bot.register_next_step_handler(message, th_check_answer)



def th_check_answer(message):
    username = message.from_user.username
    team = profiles[username]["team"]
    th_remaining = team.get_th_remaining()
    no = message.text
    int_no = int(no)
    profiles[username]["station_no"] = int_no
    remove_keyboard = types.ReplyKeyboardRemove()


    #send the clue from the treasure hunt data
    bot.send_message(message.from_user.id, "Placeholder for clue. Enter anything, it will be deemed correct", reply_markup=remove_keyboard)
    bot.register_next_step_handler(message, th_give_points)

def th_give_points(message):
    #get the data first
    username = message.from_user.username
    int_no = profiles[username]["station_no"]
    ans =  message.text
    correct_ans = ans
    team = profiles[username]["team"]

    if ans == correct_ans:   
        team.th_station_clear(int_no)
        bot.send_message(message.from_user.id, "Congrats, that is the correct answer! To send the next answer at any time, press /th_send_answer")

    else:
        bot.send_message(message.from_user.id, "Sorry, that is not the right answer! To try again, press /th_send_answer")
        

    

    
    


#Admin related stuff

@bot.message_handler(commands=["admin_functions"])

def show_admin_functions(message):
    msg = "Here are a list of commands that can be used by admins:\n\n"

    msg += "/checkadmin\n"
    msg += "/setadmin\n"
    msg += "/broadcast\n"
    msg += "/reset\n"
    msg += "/admin_start_bingo\n"
    msg += "/admin_start_supper\n"

    bot.send_message(message.from_user.id, msg)

@bot.message_handler(commands=['setadmin'])
def set_admin(message):

    #this action can only be done by admins, and requires a password before it can be done
    #password is 1234 for now

    #if the user is not an admin end the program here
    username = message.from_user.username
    if not is_admin(username):
        say_no()


    #ask the user for the password first
    bot.send_message(message.from_user.id, "Please enter the password for this action")
    bot.register_next_step_handler(message, check_pw)

def check_pw(message):
    password = "1234"

    if message.text == password:
        bot.send_message(message.from_user.id, "Please enter the telegram handle of the user you want to set as admin, without the '@'")
        bot.register_next_step_handler(message, set_admin_in_stone)

    elif message.text == 0:
        bot.send_message(message.from_user.id, "Noted. Exiting interface.")

    else:
        bot.send_message(message.from_user.id, "Please enter the correct password. To end this action, enter 0")
        bot.register_next_step_handler(message, check_pw)

def set_admin_in_stone(message):
    new_admin = message.text

    admins.append(new_admin)
    bot.send_message(message.from_user.id, "Done. The user now has admin rights.")
    

@bot.message_handler(commands=['checkuseradmin'])

def call_check_user_admin(message):
    bot.send_message(message.from_user.id, "Please enter the telegram handle of the user you want to check, without the '@'")
    bot.register_next_step_handler(message, check_user_admin)


def check_user_admin(message):
    username = message.text
    if is_admin(username):
        bot.send_message(message.from_user.id, "Yep, this user is an admin!")

    else:
        bot.send_message(message.from_user.id, "Nope, this user is not an admin :(")
    
    
#function to allow user to broadcast messages to [all] who started the bot

@bot.message_handler(commands=['broadcast'])

def call_broadcast_message(message):
    username = message.from_user.username
    if not is_admin(username):
        bot.send_message(message.from_user.id, "Sorry! This function is for admins only :(")

    else:
        bot.send_message(message.from_user.id, "You are now broadcasting to all users registered on this channel. Go ahead to type out your message which would be sent out to all users")
        bot.register_next_step_handler(message, broadcast_message)

def broadcast_message(message):
    #get user data first
    fullname = message.from_user.full_name
    username = message.from_user.username
    
    #if 0, broadcast ends here
    if message.text == "0":
        bot.send_message(message.from_user.id, "Thanks. The broadcast has ended")
        return

    if message.content_type == 'text':
        #take message from this admin, and send it out to users
        to_send = "Broadcast message from " + fullname + " @" + username + ":\n\n"
        to_send += message.text
        print(started_users)

        #broadcast the message to all started users 
        tell_everyone(to_send, 'text')

    elif message.content_type == 'photo':
        #define the photo
        photo_id = message.photo[-1].file_id

        #tell the user who this broadcast is from
        to_send = "Broadcast message from " + fullname + " @" + username + ":\n\n"
        tell_everyone(to_send, 'text')
    
        # Forward the photo to started_users
        tell_everyone(photo_id, 'photo')

    elif message.content_type == "video":
        #define the video
        video_id = message.video.file_id

        #send out
        #tell the user who this broadcast is from
        to_send = "Broadcast message from " + fullname + " @" + username + ":\n\n"
        tell_everyone(to_send, 'text')

        #send video
        tell_everyone(video_id, "video")


    bot.send_message(message.from_user.id, "Message was successfully broadcast to all users. If there is any more message, send it here below. Otherwise, enter 0 to end this function")
    bot.register_next_step_handler(message, broadcast_message)



@bot.message_handler(commands=['excel_save'])

def excel_save(message):
    username = message.from_user.username
    if not is_admin(username):
        say_no()

    else:
        bot.send_message(message.from_user.id, "Saving...")
        start = time.time()
        
        save_files("adhoc")
        end = time.time()

        time_taken = str(end-start)[:3]
        bot.send_message(message.from_user.id, "Data successfully saved to excel! (" + time_taken + "s)" )



#useful helper functions for this bot

def form_teams(teams_dict):   
    msg = ""  #empty string
    teams_list = list(teams_dict.values())
    teams_list.sort(key=lambda team: team.get_points(), reverse=True)

    for team in teams_list:
        msg += team.get_team_name()
        msg += " "
        msg += str(team.get_points())
        msg += " points"
        msg += "\n"

    return msg

def form_stations(stations_dict):   
    msg = ""  #empty string

    for stn in stations_dict.values():
        msg += stn.get_station_name()
        msg += "\n"

    return msg

def is_admin(username):
    return username in admins

def show_buttons(typ):
    if typ == "team":
        lst = teams.keys()

    elif typ == "station":
        lst = stations.keys()

    #make the keyboard
    keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)    

    #add the buttons to the keyboard
    for item in lst:
        button = types.KeyboardButton(item)
        keyboard.add(button)

    return keyboard
        

#function to output message when non-admins try to perform admin functions
def say_no():
    bot.send_message(message.from_user.id, "Sorry! This action can only be performed by admins")

#for messages to be sent to all started users
def tell_everyone(message, content_type, except_user=None):
    if content_type == 'text':
        for user_id in started_users:
            if not user_id == except_user:
                bot.send_message(user_id, message) 

    elif content_type == "photo":
        for user_id in started_users:
            if not user_id == except_user:
                bot.send_photo(user_id, message) 

    elif content_type == "video":
        for user_id in started_users:
            if not user_id == except_user:
                bot.send_video(user_id, message)


#the no more keyboard function
def remove_keyboard():
    return types.ReplyKeyboardRemove()
    
    



#class objects required in this game bot


####second thread to monitor the time and send a message every 10min

def form_time():
    time = datetime.now()

    hh = time.hour 
    mm = time.minute

    time_24h = str(hh) if hh >= 10 else "0" + str(hh)
    time_24h += str(mm) if mm >= 10 else "0" + str(mm)

    return time_24h

def form_date():
    time = datetime.now()

    day = time.day
    month = time.month

    return str(day) + "/" + str(month)


#an infinity polling function to save the data to an excel sheet every 30s

#helper function to do the saving

def save_files(typ):
    #load the template file first to work on it
    wb = load_workbook(filename="C:/Users/matth/OneDrive - National University of Singapore/Academic Files/Year 2/Orientation/template.xlsx")   #need to load workbook

    sheet = wb["Overview"]

    #get the information about time
    time = form_time()   #time is read as a string

    date = form_date()

    sheet["G5"] = time
    sheet["G4"] = date


    #count the number of players in the game
    sheet["B6"] = len(profiles.keys())   #count how many people there are in the bot


    #save the data for the teams

    r = 2
    for team in teams.values():
        team_name = team.get_team_name()
        points = team.get_points()
        
        sheet = wb["Team Information"]

        sheet.cell(row=r, column=1).value = r - 1
        sheet.cell(row=r, column=2).value = team_name
        sheet.cell(row=r, column=5).value = points

        r += 1

    #save data for each specific team (can be re-loaded later)

    for team in teams:
        team_name = team.get_team_name()
        sheet = wb[team_name]   #the sheet for this team is exactly the team name


        r = 2

        sheet.cell(row=r, column=2).value = team.get_team_name() 

        c = 3

        #add member data for as many members as we have 
         
        

    #save the file data
    filename = "C:/Users/matth/OneDrive - National University of Singapore/Academic Files/Year 2/Orientation/"
    filename += typ
    filename += "_"
    filename += time
    filename += ".xlsx"
    
    wb.save(filename)
    

def infinity_save():
    count = 0

    while True:
        time = datetime.now()
        ss = time.second
        mm = time.minute

        #does a save every minute
        if mm == 0 and ss == 0:   #sends once at the top of the min
            save_files("schd")
            bot.send_message("496804326", "files saved! this is a scheduled save every hour")
            continue 


def infinity_update():
    while True:
        time = datetime.now()
        ss = time.second
        mm = time.minute

        #does a save every minute
        if mm in [30, 0] and ss == 0:   #sends once every 30min
            no = len(teams)

            msg = "Hello! There are currently " + str(no) + " teams in the game. They are:\n"
            msg += form_teams(teams)   #convert the list of teams to word outputable format

            bot.send_message(tracking_id, msg)



@bot.message_handler(commands=["side_quest"])

def start_side_quest(message):
    #data
    username = message.from_user.username
    team = profiles[username]["team"]
    team_name = team.get_team_name()

    #check if anyone else in the team already started. If they did do not let another one start 
    if team.is_playing_side_quest():
        bot.send_message(message.from_user.id, "Sorry! Your teammate has already started a session of side quest, please send there instead")
        return

    #carry on if there is no one else playing
    team.set_playing_side_quest(True)

    #custom keyboard for the available side quests
    keyboard = generate_keyboard(*(list(team.get_side_quests().keys()) + ["Cancel"]))  #make a keyboard with the available side quests

    msg = "Welcome to side quests!"
    msg += "\n\nSide Quests completed: " + team.get_remaining_side_quests()
    msg += "\n\nWhich side quest would you like to send in?\n(To cancel the operation, press Cancel at the bottom)"

    bot.send_message(message.from_user.id, msg, reply_markup=keyboard)
    bot.register_next_step_handler(message, side_quest)

def side_quest(message):
    #check for cancel (if cancel end here)
    if message.text == "Cancel":
        username = message.from_user.username
        team = profiles[username]["team"]
        team.set_playing_side_quest(False)  #end the operation

        bot.send_message(message.from_user.id, "Ok, this process has been aborted. You may start side quests again at any time by pressing /side_quest", reply_markup=remove_keyboard())
        return 
    
    #check which side quest he is sending in for
    sq = message.text  
    username = message.from_user.username
    profiles[username]["sq"] = sq
    sq_description = all_side_quests[sq][2]   #stored at position 2 on the tup  


    #send message to the group for approval (and remove custom keyboard from prev step)
    msg = "<u>" + sq + "</u>\n\n"
    msg += "Side Quest Description: " + sq_description 
    msg += "\n\nPlease go ahead to send the photo / video for this side quest as outlined above to be reviewed by the admins"

    bot.send_message(message.from_user.id, msg, reply_markup=remove_keyboard(), parse_mode="HTML")
    bot.register_next_step_handler(message, side_quest_photo_video)


def side_quest_photo_video(message):

    #get data
    username = message.from_user.username
    sq = profiles[username]["sq"]
    team = profiles[username]["team"]
    team_name = team.get_team_name()
    sender_id = message.from_user.id   #get the chat id of the sender to get back to them later

    #1.to the exco group, send the opener first (tell them which team sent in and which sq they are attempting)
    first_msg = "Submission from " + team_name + ":"
    first_msg += "\nSide Quest: " + sq

    writeup = all_side_quests[sq][2]
    first_msg += "\n\nDescription of side quest: " + writeup 

    bot.send_message(exco_id, first_msg)

        
    #2. process what the user sent to the bot

    if message.content_type == 'text':
        #if just a text, simple just send it over to the admin chat
        to_send = message.text

        #broadcast the message to all started users 
        bot.send_message(exco_id, to_send)

    elif message.content_type == 'photo':
        #define the photo
        photo_id = message.photo[-1].file_id
        
        # Send the photo to the exco chat
        bot.send_photo(exco_id, photo_id)

    elif message.content_type == "video":
        #define the video
        video_id = message.video.file_id

        #send video
        bot.send_video(exco_id, video_id)
        
        #to include support for videos to be sent to the bot later


    #3. close with the last msg 
    last_msg = "To approve / reject side quest answers at any time, press /approve_side_quest"
    msg = bot.send_message(exco_id, last_msg)
    msg_id = msg.message_id
    
    #4. save this info to the log first to be retrieved later 
    active_side_quests[team_name + ", " + sq] = (sq, team_name, msg_id, sender_id)   #key is team_name, sq, and the value is the id of the msg and sender

    #5. then at the end tell the user that the photo / text was sent out successfully
    bot.send_message(message.from_user.id, "Your photo/video has been successfully sent! You will be notifed again after an \
admin reviews your submission")
                     
    #6. Set their playing side quest status to False 
    team.set_playing_side_quest(False)



#approve side quest and give out wildcards function, only to be used by admins
@bot.message_handler(commands=["approve_side_quest"])

def call_approve_side_quest(message):
    username = message.from_user.username
    profiles[username]["to_delete_sq"] = []
    profiles[username]["to_delete_sq"].append(message.message_id)

    #get the data of the submitted side quests first
    active_sq_items = active_side_quests.keys()  #convert the list of tuples to team, sq
    print(active_sq_items)
    keyboard = generate_keyboard(*active_sq_items)

    #then sent this to the user and ask which one they want to approve

    msg = bot.reply_to(message, "Which side quest would you like to judge?", reply_markup=keyboard)

    #make a list of the messages to be deleted later 
    msg_id = msg.message_id
    profiles[username]["to_delete_sq"].append(msg_id)

    #next step
    bot.register_next_step_handler(message, side_quest_yes_or_no)

def side_quest_yes_or_no(message):
    #register which team / side quest you are looking at 
    username = message.from_user.username
    curr_sq = message.text   #the text of the side quest
    profiles[username]["curr"] = curr_sq   

    #log data of the messages to delete later on
    profiles[username]["to_delete_sq"].append(message.message_id)

    #give them the value as required

    #ask if it is a yes or no
    keyboard = generate_keyboard("Yes", "No")  #keyboard has two options
    msg = bot.reply_to(message, "Is this correct?", reply_markup=keyboard)
    msg_id = msg.message_id
    profiles[username]["to_delete_sq"].append(msg_id)

    bot.register_next_step_handler(message, side_quest_check_answer)

def side_quest_check_answer(message):
    #log data
    username = message.from_user.username
    profiles[username]["to_delete_sq"].append(message.message_id)

    #get the ans 
    ans = message.text

    #process
    if ans == "Yes":
        #data
        full_name = message.from_user.full_name
        sq_text = profiles[username]["curr"]
        sq_name, team_name, msg_id, sender_id = active_side_quests[sq_text] 
        team = teams[team_name]
        wildcard, colour, description = all_side_quests[sq_name] #since the tup is card name, colour
        effect = wild_cards[wildcard]

        #1. tell the team they got it
        msg = "Congrats! Your submission for " + sq_name + " is correct and you have received the <b>" + colour + " </b>wildcard " + str(wildcard)
        msg += "\n\n" + str(wildcard) + ": " + effect
        msg += "\n\nSide Quests completed: " + str(team.get_remaining_side_quests(1))  #1 to account for the one done here
        bot.send_message(sender_id, msg, parse_mode="HTML")
  
        #2. add the wildcard to the team and remove from team list (and effects if any)
        if colour == "Green":
            if wildcard == "Dice Roll":
                greenwc_dice_roll(message, sender_id, team)

            else:
                team.add_green_wildcard(wildcard, effect)

        elif colour == "Red":
            #no need to add red wildcard for og and expose since they are immediate effect
            if wildcard == "Own Goal":
                redwc_own_goal(message, sender_id)

            elif wildcard == "Expose":
                redwc_expose(message)

            elif wildcard == "Dementia":
                redwc_dementia(message, sender_id)

            else:
                team.add_red_wildcard(wildcard, effect)    #add for the rest 

        #2.5 remove the side quest for this team 
        team.remove_side_quest(sq_name)
        print("side quest removed for the team")
        
        #3. edit the original message in the group to say that this has been ammended 
        print(msg_id, sender_id)
        bot.edit_message_text(chat_id=exco_id, message_id=msg_id, text="Edit: Approved by " + full_name + " @" + username + " at " + form_time() + "H")

        #eg. the message is now "Edit: Approved by Matthew Yip @matthewyip at 1000H"


        #4. tell the group it has been awarded
        bot.reply_to(message, "Great! Wildcard " + wildcard + " has been awarded to team " + team_name, reply_markup=remove_keyboard())


        #5. Update the data to remove this entry
        active_side_quests.pop(sq_text)


        #6. Delete the back and forth messages for it to be neater
        for msg_id in profiles[username]["to_delete_sq"]:
            bot.delete_message(chat_id=exco_id, message_id=msg_id)


    elif ans == "No":
        #reply msg and log the msg sent
        msg = bot.reply_to(message, "Ok, plese include a message to be sent to the OG to inform them why the submission is not correct", reply_markup=remove_keyboard())
        msg_id = msg.message_id
        profiles[username]["to_delete_sq"].append(msg_id)

        bot.register_next_step_handler(message, side_quest_tell_them_no)

    else:
        bot.reply_to(message, "Sorry, that is not a valid option! Please press the buttons for Yes or No")
        bot.register_next_step_handler(message, side_quest_check_answer)

def side_quest_tell_them_no(message):
    #log data
    username = message.from_user.username
    profiles[username]["to_delete_sq"].append(message.message_id)

    #get the text to send them 
    rejection_reason = message.text

    #get data
    username = message.from_user.username
    full_name = message.from_user.full_name
    sq_text = profiles[username]["curr"]
    sq_name, team_name, msg_id, sender_id = active_side_quests[sq_text] 
    team = teams[team_name]

    wildcard = all_side_quests[sq_name][0]  #since the tup is card name, colour
    
    #1. Tell the team that the submission is not correct
    bot.send_message(sender_id, "Hey, your submission for side quest " + sq_name + " was not correct :( Here's a message from the admins:\n\n" + rejection_reason +"\n\nTo re-submit at any time, press /side_quest")


    #2. Edit the message to show that it is not correct
    new_text = "Edit: Rejected by " + full_name + " @" + username + " at " + form_time() + "H\n\n"
    new_text += "Reason: " + rejection_reason
    bot.edit_message_text(chat_id=exco_id, message_id=msg_id, text=new_text)


    #3. Send in the exco chat that it is rejected
    bot.send_message(exco_id, "Got it, informed the team that it is not correct")

    #4. Remove information from active, and set the team to not playing 
    active_side_quests.pop(sq_text)

    #5. Delete all texts for neatness
    for msg_id in profiles[username]["to_delete_sq"]:
        bot.delete_message(chat_id=exco_id, message_id=msg_id)



#######FUNCTIONS FOR THE WILDCARD EFFECTS FOR "IMMEDIATE EFFECT" WILDCARDS - BOTH RED AND GREEN#################

def greenwc_dice_roll(message, team_id, team):
    bot.send_message(team_id, "Great! Your team got the green wild card Dice Roll which will now roll a dice (via python random module), and you will get the number of the dice x 10 points added to your team!")
    bot.send_message(team_id, "Rolling Dice...")

    time.sleep(3)

    dice_roll = random.randint(1, 6)   #simulates dice roll
    pts_add = dice_roll * 10 

    team.add_points(pts_add, "Side Quest Green Wildcard - Dice Roll")

    bot.send_message(team_id, "Results are in! Your team rolled a \n\n" + str(dice_roll) + "\n\nYour team has gained " + str(pts_add) + " points!")


def redwc_own_goal(message, team_id):
    #data
    username = message.from_user.username
    team = profiles[username]["team"]
    team_name = team.get_team_name()


    #1. Dock 10 points from this team
    team.add_points(-10, "Red Wildcard - Own Goal")   #the 10 points is lost from "Red Wildcard - Own Goal"

    #2. Tell the team of their fate 
    bot.send_message(team_id, "Oops! Your team got the red wild card Own Goal and just lost 10 points..")


def redwc_expose(message):
    #data
    username = message.from_user.username
    team = profiles[username]["team"]
    team_name = team.get_team_name()


    #1. broadcast out what green wildcards this team has
    msg = "Broadcast message from Side Quest Admin:\n\n"
    msg += "Attention! Attention! Team " + team_name + " got hit with the red wildcard Expose, and will now have its list of green wildcards exposed to everyone."
    msg += " It is holding the following wildcards:\n\n"
    
    green_wc = list(team.get_green_wildcards().keys())

    if green_wc == []:
        msg += "NIL"

    for wc in green_wc:
        msg += wc
        msg += "\n"

    tell_everyone(msg, "text")


    #2. tell the team that they got exposed
    bot.send_message(message.from_user.id, "Oops, your team got hit with the red wildcard Expose and your list of green wildcards has been exposed to everyone to see....")


def redwc_dementia(message, sender_id):
    #data
    username = message.from_user.username
    team = profiles[username]["team"]
    team_name = team.get_team_name()

    #1. do the action 
    lost_wc = team.your_loss()

    #2. tell the team
    bot.send_message(sender_id, "Oops, your team got hit with the red wildcard Your Loss and you have lost your most recently obtained green wildcard which is:\n\n" + lost_wc)


#this function is used to convert all red cards at the end to -20 points per card

@bot.message_handler(commands=["admin_convert_red_cards"])

def call_admin_convert_red_cards(message):
    #this function can only be invoked by admins!
    if not is_admin(message.from_user.username):
        say_no()
        return
    
    keyboard = generate_keyboard("Yes", "No")
    
    #if not confirm first from the user that they want to do so
    bot.send_message(message.from_user.id, "This action would convert all red cards to 20 points lost per card the team is still holding. Are you sure you want to proceed with this action? (Should only be done immediately before the commencement of supper bidding)", reply_markup=keyboard)
    bot.register_next_step_handler(message, convert_red_cards)

def convert_red_cards(message):
    #if no abort the process
    if message.text == "No":
        bot.send_message(message.from_user.id, "Ok, this action has been cancelled", reply_markup=remove_keyboard()) 
        return #end the process here
    
    #if not proceed with the action
    #1. Carry out the action to give -20 points per team per red card holding
    for team in list(teams.values()):
        team.convert_red_cards()

    #2. tell the teams of this action
    msg = "Broadcast message from Side Quest admin:\n\n"
    msg += "And that is the end of the side quests, time to start the supper bidding! For teams still holding on to red cards, oops.. you have lost 20 points per red card still holding on to :("

    tell_everyone(msg, "text")

    #3. Tell the admin who triggered this process that it has been successfully carried out
    bot.send_message(message.from_user.id, "Ok, this action has been carried out successfully! Each team has lost 20 points for each red wildcard they are still holding on to.", reply_markup=remove_keyboard())


@bot.message_handler(commands=["admin_use_wildcard"])

def admin_use_wildcard(message):
    #check user first (ONLY matthew yip can use this function)
    username = message.from_user.username
    if username != "matthewyip":
        say_no()

    keyboard = generate_keyboard(*list(teams.keys()))
    
    bot.send_message(message.from_user.id, "Ok, which team used a wildcard?", reply_markup=keyboard)
    bot.register_next_step_handler(message, check_wildcard)

def check_wildcard(message):
    #get data
    team_name = message.text
    username = message.from_user.username
    profiles[username]["curr"] = team_name
    team = teams[team_name]
    wildcards = list(team.get_green_wildcards().keys())

    #check if this team has any wildcards first 
    if wildcards == []:
        bot.send_message(message.from_user.id, "This team has no wildcards!", reply_markup=remove_keyboard())
        return 
    
    #if they do and its valid then go ahead to apply
    keyboard = generate_keyboard(*wildcards)

    bot.send_message(message.from_user.id, "Ok, what wildcard did they use?", reply_markup=keyboard)
    bot.register_next_step_handler(message, apply_wildcard)

def apply_wildcard(message):
    username = message.from_user.username
    team_name = profiles[username]["curr"]
    team = teams[team_name]
    wildcard = message.text

    team.remove_green_wildcard(wildcard)
    bot.send_message(message.from_user.id, "Ok, wildcard " + wildcard + " has been consumed for the team " + team_name, reply_markup=remove_keyboard())



#start the thread for saving information
save_requests = threading.Thread(target=infinity_update)
save_requests.start()
                                                                                                                                 

#to run the bot on an infinite loop
bot.infinity_polling()
